#!/usr/bin/env python3
"""Build a golden-hour event schedule spreadsheet from cameras.json.

Tabs: sunrises, sunsets, all_events.
Times are shown in the viewer's timezone (default America/Chicago) and UTC.
"""
import json, math, sys
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HOME_TZ = ZoneInfo(sys.argv[2] if len(sys.argv) > 2 else 'America/Chicago')
GOLDEN_DEG = 6.0


def solar_elev(lat, lon, dt):
    """Sun elevation in degrees, and whether it is climbing, at a UTC datetime."""
    rad = math.pi / 180
    doy = (dt - datetime(dt.year, 1, 1, tzinfo=timezone.utc)).total_seconds() / 86400
    h = dt.hour + dt.minute / 60 + dt.second / 3600
    g = 2 * math.pi / 365 * (doy + (h - 12) / 24)
    eq = 229.18 * (0.000075 + 0.001868 * math.cos(g) - 0.032077 * math.sin(g)
                   - 0.014615 * math.cos(2 * g) - 0.040849 * math.sin(2 * g))
    decl = (0.006918 - 0.399912 * math.cos(g) + 0.070257 * math.sin(g)
            - 0.006758 * math.cos(2 * g) + 0.000907 * math.sin(2 * g)
            - 0.002697 * math.cos(3 * g) + 0.00148 * math.sin(3 * g))
    tst = h * 60 + eq + 4 * lon
    ha = (tst / 4) - 180
    while ha < -180: ha += 360
    while ha > 180:  ha -= 360
    cz = (math.sin(lat * rad) * math.sin(decl) +
          math.cos(lat * rad) * math.cos(decl) * math.cos(ha * rad))
    cz = max(-1.0, min(1.0, cz))
    return 90 - math.acos(cz) / rad, ha < 0


def scan(cam, start_utc, minutes):
    """Return golden-window events for one camera over the scan range."""
    out, cur, prev = [], None, None
    for m in range(minutes + 1):
        t = start_utc + timedelta(minutes=m)
        elev, rising = solar_elev(cam['lat'], cam['lon'], t)
        inw = abs(elev) <= GOLDEN_DEG
        if inw and cur is None:
            cur = {'rising': rising, 'start': t, 'cross': None}
        if cur is not None:
            if prev is not None and cur['cross'] is None and (prev <= 0 < elev or prev >= 0 > elev):
                cur['cross'] = t          # actual horizon crossing
            if not inw:
                cur['end'] = t
                out.append(cur)
                cur = None
        prev = elev
    if cur is not None:
        cur['end'] = start_utc + timedelta(minutes=minutes)
        out.append(cur)
    return [e for e in out if e['cross'] is not None]


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'cameras.json'
    data = json.load(open(path))
    cams = [c for c in data['cameras'] if isinstance(c.get('lat'), (int, float))]

    # Cover the next full local day, starting from local midnight today.
    now_home = datetime.now(HOME_TZ)
    start_home = now_home.replace(hour=0, minute=0, second=0, microsecond=0)
    start_utc = start_home.astimezone(timezone.utc)
    MINUTES = 24 * 60

    events = []
    for c in cams:
        for e in scan(c, start_utc, MINUTES):
            events.append({
                'cam': c, 'rising': e['rising'],
                'cross': e['cross'], 'start': e['start'], 'end': e['end'],
            })
    events.sort(key=lambda e: e['cross'])

    # concurrency: how many other golden windows are open at each crossing
    for e in events:
        e['concurrent'] = sum(1 for o in events
                              if o is not e and o['start'] <= e['cross'] < o['end'])

    HEAD = ['Event', 'Camera', 'Where', 'State', 'Feed',
            'Event time (%s)' % start_home.tzname(), 'Event time (UTC)',
            'Golden starts', 'Golden ends', 'Window (min)',
            'Other cams live then', 'Lat', 'Lon', 'Categories']

    def row(e):
        c = e['cam']
        loc = lambda t: t.astimezone(HOME_TZ).strftime('%H:%M')
        return ['Sunrise' if e['rising'] else 'Sunset',
                c['name'], c.get('sub', ''), c.get('state', ''),
                {'hls': 'live video', 'yt': 'live video', 'jpg': 'snapshot'}.get(c['type'], c['type']),
                loc(e['cross']), e['cross'].strftime('%H:%M'),
                loc(e['start']), loc(e['end']),
                int(round((e['end'] - e['start']).total_seconds() / 60)),
                e['concurrent'], c['lat'], c['lon'],
                ', '.join(t for t in c.get('tags', []) if t != 'golden')]

    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0B1220')
    head_font = Font(bold=True, color='FFFFFF')
    rise_fill = PatternFill('solid', fgColor='FFF4CC')
    set_fill  = PatternFill('solid', fgColor='FFE3D1')

    def sheet(title, rows, tint):
        ws = wb.create_sheet(title)
        ws.append(HEAD)
        for cell in ws[1]:
            cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        if tint:
            for rr in ws.iter_rows(min_row=2, max_col=1):
                for cell in rr:
                    cell.fill = rise_fill if cell.value == 'Sunrise' else set_fill
        widths = [9, 34, 34, 6, 11, 15, 15, 13, 12, 12, 18, 9, 10, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        return ws

    wb.remove(wb.active)
    sheet('sunrises',   [row(e) for e in events if e['rising']],      True)
    sheet('sunsets',    [row(e) for e in events if not e['rising']],  True)
    sheet('all_events', [row(e) for e in events],                     True)

    out = sys.argv[3] if len(sys.argv) > 3 else 'golden_hour_schedule.xlsx'
    wb.save(out)
    rises = sum(1 for e in events if e['rising'])
    print('%s — %d events (%d sunrises, %d sunsets) for %s'
          % (out, len(events), rises, len(events) - rises, start_home.strftime('%a %Y-%m-%d')))
    peak = max(e['concurrent'] for e in events) if events else 0
    print('peak simultaneous golden windows: %d' % (peak + 1))


if __name__ == '__main__':
    main()
